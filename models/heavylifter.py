from openpyxl import load_workbook
import pandas as pd
from datetime import datetime as dt
import os


class HeavyLifter:
    """
    This is the main bulk of script.
    It takes in the list of files to be extracted and does the extracting
    """

    def __init__(self, target_dir, target_files, scraper_df, server_conn_string):
        # __init__ variables
        self.target_dir = target_dir
        self.target_files = target_files
        self.scraper_df = scraper_df
        self.server_conn_string = server_conn_string
        # executing methods
        self.run()

    def run(self):
        # change to dir with workbooks
        os.chdir(self.target_dir)

        # TODO: implement try/excepts

        # iterate over all workbooks
        for file in self.target_files:
            # load workbook
            wb = load_workbook(file, read_only=True, data_only=True)

            # working hierarchically, iterate over all tables, columns, cells and create dataframe
            for sql_table in self.scraper_df['sql_table'].unique():
                # filter df so only looking at this particular SQL table
                df_filt_sql = self.scraper_df[self.scraper_df['sql_table'] == sql_table]

                # list of worksheets to grab data to put into this sql table
                worksheets = df_filt_sql['source_sheet'].unique()
                for sheet in worksheets:
                    ws = wb[sheet]

                    # data dict for use in creating df later
                    data_dict = dict()
                    data_types_dict = dict()

                    for column_name in df_filt_sql['column_name']:

                        # filtering for the row
                        df_filt_col = df_filt_sql[df_filt_sql['column_name'] == column_name]

                        # extracting row values for use in extraction and data-typing later
                        range_or_cell = df_filt_col['is_target_a_cell_or_range_of_cells'].values[0]
                        range_type = df_filt_col['range_type'].values[0]
                        target_cell = df_filt_col['target_cell'].values[0]
                        target_range = df_filt_col['target_range'].values[0]

                        # making record of data type for casting in pandas later
                        data_types_dict[column_name] = df_filt_col['sql_data_type'].values[0]

                        print("Extraction beginning")
                        if range_or_cell == 'cell':
                            print("Cell")
                            data_dict[column_name] = ws[target_cell].value
                            print("Cell Done.\n")
                        elif range_or_cell == 'range':
                            print("Range.")
                            if range_type == 'row':
                                #  iterate over each row in the target row and append value
                                # TODO: Examine more efficient method of row iteration
                                print("Row.")
                                print(sql_table, sheet, column_name, target_range)
                                data_dict[column_name] = [str(cell.value) for cell in ws[target_range][0]]
                                print("Row done.")
                            elif range_type == 'column':
                                print("Col.")
                                #  iterate over each cell in the target column and append value
                                data_dict[column_name] = []
                                for cell in ws[target_range]:
                                    data_dict[column_name].append(cell[0].value)
                                print("Col done.")
                            print("Range done.\n")
                    # adding in file and upload info
                    data_dict['file_name'] = file
                    data_dict['upload_utc'] = dt.utcnow()

                    # send to SQL server
                    # try and except to handle if array is present in dict
                    try:
                        df = pd.DataFrame(data=data_dict, dtype='object')
                    except ValueError:
                        df = pd.DataFrame(data=data_dict, dtype='object', index=[0])


                    # TODO implement proper switch between SQL and CSV
                    df.to_sql(name=sql_table,
                              if_exists='append',
                              con=self.server_conn_string)
